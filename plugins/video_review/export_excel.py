"""Excel export utilities for the Video Review & Log plugin."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .model import VideoItem

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD = Font(bold=True)


def _format_duration(seconds: int | None) -> str:
    if seconds is None or seconds < 0:
        return "--"
    minutes, sec = divmod(seconds, 60)
    return f"{int(minutes):02d}:{int(sec):02d}"


def export_to_excel(path: Path, items: Iterable[VideoItem]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    headers = [
        "Index",
        "Filename",
        "Relative Path",
        "Size (MB)",
        "Modified (ISO)",
        "Duration",
        "Status",
        "Notes",
    ]
    ws.append(headers)

    for row_idx, item in enumerate(items, start=2):
        size_mb = item.size_bytes / (1024 * 1024)
        modified = datetime.fromtimestamp(item.mtime).isoformat()
        duration = _format_duration(item.duration_seconds)
        status_text = "Pending"
        fill = None
        font = None
        if item.status == "approved":
            status_text = "✓ Approved"
            fill = GREEN_FILL
            font = BOLD
        elif item.status == "rejected":
            status_text = "✗ Rejected"
            fill = RED_FILL
            font = BOLD
        row = [
            item.index,
            item.filename,
            item.rel_path,
            round(size_mb, 2),
            modified,
            duration,
            status_text,
            item.note,
        ]
        ws.append(row)
        status_cell = ws.cell(row=row_idx, column=7)
        if fill:
            status_cell.fill = fill
        if font:
            status_cell.font = font

    # Freeze header row and add filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto-fit columns based on content
    for column_idx, column_title in enumerate(headers, start=1):
        column_letter = get_column_letter(column_idx)
        max_length = len(column_title)
        for cell in ws[column_letter]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(Path(path))
